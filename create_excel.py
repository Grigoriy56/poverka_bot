import pandas
import openpyxl
from openpyxl.styles import PatternFill, Font
from random import randint
from data import counter
from datetime import timedelta, date

# создание xlsx файла
# создание xlsx файла


def create_excel(number):
    wb = openpyxl.Workbook()
    sheet = wb.active
    stolb_name = ['TypePOV', 'GosNumberPOV', 'NamePOV', 'DesignationSiPOV', 'DeviceMarkPOV', 'type',
           'SerialNumPOV', 'InventarNumPOV', 'CalibrationDatePOV', 'NextcheckDatePOV', 'MarkCipherPOV',
           'DocPOV', 'DeprcatedPOV', 'temperature', 'pressure', 'hymidity', 'other', 'StandartPOV',
           'GpsPOV', 'SiPOV', 'SiPOVetalon', 'SoPOV', 'ReagentNumber', 'ManufactureYear', 'isOwner',
           'gpsTitle', 'lpsTitle', 'npeNumber', 'rank', 'mpTitle', 'regNumber', 'miOwner', 'signPass',
           'signMi', 'metrologist', 'oMethod', 'calibration', 'reasons', 'structure', 'characteristics',
           'additional_info', 'mimetype', 'filename']
    count = -1
    redFill = PatternFill(start_color='F5F5DC',
                          end_color='F5F5DC',
                          fill_type='solid')
    for row in sheet['A1':'AQ1']:
        for cell in row:
            count += 1
            cell.value = stolb_name[count]
            cell.fill = redFill
            cell.font = Font(size='8', name='Arial')
    wb.save(f'poverka/poverka{number}.xlsx')


# mas хранит занчение ячеек, например mas = [1, 2, 3 ...], значит в ячейке Ax = 1, Bx = 2
# , где x равен номеру строчки
# date хранит в себе номер строчки, куда нужно добавлять значение
# number хранит в себе номер xlsx файла
def add_line(mas, number):
    wb = openpyxl.load_workbook(filename=f'poverka/poverka{number}.xlsx')
    excel_data_df = pandas.read_excel(f'poverka/poverka{number}.xlsx')
    number_line = len(excel_data_df.to_dict(orient='record'))
    sheet = wb.active
    count = 0
    for row in sheet[f'A{number_line+2}':f'AQ{number_line+2}']:
        for cell in row:
            cell.value = mas[count]
            cell.font = Font(size='8', name='Arial')
            count += 1
    ws = sheet
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    wb.save(filename=f'poverka/poverka{number}.xlsx')
    wb.close()


def preparation(gos_num, temp_n, data_n):
    def random_temp(x):
        if x == 1:
            return randint(56, 62)
        if x == 2:
            return randint(9, 11)
    final = [1, gos_num, '', '', counter[gos_num][0], 0, 'serial_number', 'invent', data_n,
             data_n + timedelta(days=counter[gos_num][temp_n]), '',
             '"документом МИ 1592-15 "Рекомендация. ГСИ. Счетчики воды. Методика поверки"', 'Пригодно',
             '24,0 °С', '102,0 кПа', '54,0%', f'Температура поверочной жидкости {random_temp(temp_n)} °С', '', '', '',
             '40391.09.3Р.00139988', '', '', '', '', '', '', '', '', '', '', 'ФЛ', 0, 0, 'Нестеров Е.В.', '', '', '',
             '', '', '', '', '']

    return final

